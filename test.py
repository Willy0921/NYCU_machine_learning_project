import numpy as np
import pickle
import pandas as pd
import openpyxl
from build_decision_tree import *
from argparse import ArgumentParser

parser = ArgumentParser()
parser.add_argument("-d", "--dataset", help="Dataset source", type=str, default="dataset1")
args = parser.parse_args()


def traverse_tree(instance, node):
    if node.partition_feature_idx == None:
        label_set = np.unique(node.dataset.y_train)
        y_occurance = np.array([np.sum(node.dataset.y_train == label) for label in label_set])
        return label_set[np.argmax(y_occurance)]
    else:
        if instance[node.partition_feature_idx] < node.partition_feature_val:
            node = node.children[0]
            return traverse_tree(instance, node)
        else:
            node = node.children[1]
            return traverse_tree(instance, node)


def predict_target_value(instance, tree_node):
    return traverse_tree(instance, tree_node)


def load_dataset():

    X_train = pd.read_excel(f"./{args.dataset}/X_test_processed.xlsx").to_numpy()
    return X_train


def save_prediction(y_predict):
    y_predict = y_predict.tolist()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.dataset
    if args.dataset == "dataset1":
        ws.cell(row=1, column=1).value = "class"
    elif args.dataset == "dataset2":
        ws.cell(row=1, column=1).value = "Sentiment"
    for instance in y_predict:
        ws.append(instance)
    wb.save(filename=f"./prediction/{args.dataset}/y_predict.xlsx")
    print("Saving prediction done!")


if __name__ == "__main__":
    X_test = load_dataset()

    if args.dataset == "dataset1":
        with open("./model/decision_tree_model1.pkl", "rb") as f:
            root = pickle.load(f)
    elif args.dataset == "dataset2":
        with open("./model/decision_tree_model2.pkl", "rb") as f:
            root = pickle.load(f)

    y_predict = np.array([predict_target_value(X_test[instance, :], root) for instance in range(X_test.shape[0])])
    y_predict = y_predict.reshape((y_predict.shape[0], -1))
    save_prediction(y_predict)
